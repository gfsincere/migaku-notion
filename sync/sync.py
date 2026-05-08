#!/usr/bin/env python3
"""Pull words from a running migoku server and upsert them into a Notion database.

Migoku exposes Migaku's local SRS database as a REST API on http://localhost:8080.
This script:
  1. Authenticates with migoku (using email+password from .env, or a pre-derived API key)
  2. Fetches all words for the chosen language (paginated)
  3. Fetches the difficult-words endpoint to enrich rows with fail rate / review counts
  4. Diffs each word against a local SQLite cache of {migaku_key -> notion_page_id +
     tracked fields}, and only PATCHes Notion rows whose tracked fields actually
     changed. New words get created and the cache is updated atomically right after
     each successful Notion call.

The local cache lives at ``sync/state.db`` (SQLite, stdlib only). Delete it and run
``python sync.py rebuild-cache`` to recreate it from Notion.

Important: the Notion "Meaning" column is NEVER overwritten on update, so any
AI-generated meanings are preserved across syncs. Only new rows get a blank Meaning.
"""

from __future__ import annotations

import argparse
import getpass
import logging
import os
import re
import sqlite3
import sys
import time
import webbrowser
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import requests
from dotenv import load_dotenv

load_dotenv()

NOTION_VERSION = "2022-06-28"
NOTION_API = "https://api.notion.com/v1"
DEFAULT_LANG = os.getenv("SYNC_LANG", "zh")

# pypinyin is optional at import time but required at run time for lang=zh syncs.
try:
    from pypinyin import lazy_pinyin, Style as _PinyinStyle  # type: ignore
    _PINYIN_AVAILABLE = True
except ImportError:
    _PINYIN_AVAILABLE = False
    _PinyinStyle = None  # type: ignore

# Local cache lives next to this script.
STATE_DB_PATH = Path(__file__).resolve().parent / "state.db"

log = logging.getLogger("migaku-notion")


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

@dataclass
class Word:
    dict_form: str
    secondary: str  # Migaku's disambiguation index (0/1/...) for zh; kana reading for ja
    known_status: str
    language: str
    fail_rate: float | None = None
    total_reviews: int | None = None
    failed_reviews: int | None = None
    part_of_speech: str | None = None
    # For zh: tone-marked pinyin (e.g. "xué xí") and numeric pinyin (e.g. "xue2 xi2"),
    # both derived from `dict_form` via pypinyin. None for non-zh.
    pinyin_marks: str | None = None
    pinyin_numeric: str | None = None

    @property
    def key(self) -> str:
        return f"{self.language}|{self.dict_form}|{self.secondary}"


@dataclass
class CachedRow:
    """One row in sync/state.db. Mirrors the `words` table 1:1."""

    migaku_key: str
    page_id: str
    lang: str
    dict_form: str
    secondary: str
    known_status: str | None
    fail_rate: float | None
    total_reviews: int | None
    failed_reviews: int | None
    part_of_speech: str | None
    last_synced: str | None
    archived: bool
    pinyin_marks: str | None = None    # Notion "Pinyin" column (tone marks)
    pinyin_numeric: str | None = None  # Notion "Pinyin (numeric)" column
    sense_index: str | None = None     # Notion "Sense #" column (mirrors `secondary` for zh)


def compute_pinyin_marks(hanzi: str) -> str:
    """Generate tone-marked pinyin (e.g. 'xué xí'). Returns "" if pypinyin unavailable."""
    if not hanzi or not _PINYIN_AVAILABLE:
        return ""
    return " ".join(lazy_pinyin(hanzi, style=_PinyinStyle.TONE))


def compute_pinyin_numeric(hanzi: str) -> str:
    """Generate numeric-tone pinyin (e.g. 'xue2 xi2'). Returns "" if pypinyin unavailable."""
    if not hanzi or not _PINYIN_AVAILABLE:
        return ""
    return " ".join(lazy_pinyin(hanzi, style=_PinyinStyle.TONE3))


# ---------------------------------------------------------------------------
# Migoku client
# ---------------------------------------------------------------------------

class MigokuClient:
    def __init__(self, base_url: str, api_key: str | None = None) -> None:
        self.base = base_url.rstrip("/")
        self.api_key = api_key
        self.session = requests.Session()

    def login(self, email: str, password: str) -> str:
        log.info("Logging in to migoku as %s ...", email)
        resp = self.session.post(
            f"{self.base}/auth/login",
            json={"email": email, "password": password},
            timeout=120,
        )
        resp.raise_for_status()
        data = resp.json()
        self.api_key = data["api_key"]
        log.info("Migoku login OK (%s)", data.get("message", ""))
        return self.api_key

    def _headers(self) -> dict[str, str]:
        if not self.api_key:
            raise RuntimeError("Not authenticated to migoku. Set MIGOKU_API_KEY or login first.")
        return {"X-Api-Key": self.api_key}

    def list_words(
        self,
        lang: str,
        page_size: int = 500,
        statuses: list[str] | None = None,
    ) -> Iterable[Word]:
        """Yield Word objects from migoku, optionally restricted to one or more statuses.

        Migoku's `/api/v1/words` only accepts a single `status` query param at a time,
        so we make one paginated call per requested status and let the caller dedupe.
        Pass `statuses=None` (or empty) to fetch all statuses in a single sweep.
        """
        for status_filter in (statuses or [None]):
            page = 1
            while True:
                params: dict[str, Any] = {"lang": lang, "page": page, "page_size": page_size}
                if status_filter:
                    params["status"] = status_filter.lower()
                resp = self.session.get(
                    f"{self.base}/api/v1/words", params=params,
                    headers=self._headers(), timeout=60,
                )
                resp.raise_for_status()
                payload = resp.json()
                for row in payload.get("data", []) or []:
                    yield Word(
                        dict_form=row.get("dictForm", ""),
                        secondary=row.get("secondary", ""),
                        known_status=row.get("knownStatus", "UNKNOWN") or "UNKNOWN",
                        language=lang,
                    )
                meta = payload.get("pagination", {})
                label = status_filter.upper() if status_filter else "ALL"
                log.info("migoku [%s] page %d/%d (%d total)", label, meta.get("page", page),
                         meta.get("total_pages", "?"), meta.get("total", "?"))
                if not meta.get("has_next"):
                    break
                page += 1

    def list_difficult_words(self, lang: str, limit: int = 2000) -> list[dict[str, Any]]:
        resp = self.session.get(
            f"{self.base}/api/v1/words/difficult",
            params={"lang": lang, "limit": limit},
            headers=self._headers(),
            timeout=60,
        )
        resp.raise_for_status()
        return resp.json() or []


# ---------------------------------------------------------------------------
# Notion client (raw HTTP, since the official SDK lags behind the API)
# ---------------------------------------------------------------------------

class NotionClient:
    """Tiny Notion API wrapper sized for our upsert workflow."""

    # Notion's documented limit is 3 requests/sec averaged. We aim for ~2.5.
    REQUEST_INTERVAL = 0.4

    def __init__(self, token: str, database_id: str) -> None:
        self.token = token
        self.database_id = database_id
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {token}",
            "Notion-Version": NOTION_VERSION,
            "Content-Type": "application/json",
        })
        self._last_call = 0.0

    def _throttle(self) -> None:
        wait = self.REQUEST_INTERVAL - (time.monotonic() - self._last_call)
        if wait > 0:
            time.sleep(wait)
        self._last_call = time.monotonic()

    def _request(self, method: str, path: str, **kw: Any) -> dict[str, Any]:
        self._throttle()
        url = f"{NOTION_API}{path}"
        for attempt in range(5):
            try:
                resp = self.session.request(method, url, timeout=60, **kw)
            except (requests.exceptions.ReadTimeout,
                    requests.exceptions.ConnectionError,
                    requests.exceptions.ChunkedEncodingError) as exc:
                # Transient network issue. Notion is generally idempotent on PATCH/POST
                # of the same payload, so retrying with backoff is safe for our use case.
                wait = 2 ** attempt
                log.warning("Notion %s on %s %s, retrying in %ds (attempt %d)",
                            type(exc).__name__, method, path, wait, attempt + 1)
                time.sleep(wait)
                continue
            if resp.status_code == 429:
                retry = float(resp.headers.get("Retry-After", "1"))
                log.warning("Notion 429, sleeping %.1fs (attempt %d)", retry, attempt + 1)
                time.sleep(retry)
                continue
            if resp.status_code >= 500:
                log.warning("Notion %d, retrying...", resp.status_code)
                time.sleep(2 ** attempt)
                continue
            if not resp.ok:
                raise RuntimeError(f"Notion {method} {path} -> {resp.status_code}: {resp.text[:500]}")
            return resp.json()
        raise RuntimeError(f"Notion {method} {path} failed after 5 attempts")

    def query_all_pages(self) -> list[dict[str, Any]]:
        """Fetch every page in the database. Returns the raw page objects."""
        pages: list[dict[str, Any]] = []
        cursor: str | None = None
        while True:
            body: dict[str, Any] = {"page_size": 100}
            if cursor:
                body["start_cursor"] = cursor
            data = self._request("POST", f"/databases/{self.database_id}/query", json=body)
            pages.extend(data.get("results", []))
            if not data.get("has_more"):
                break
            cursor = data.get("next_cursor")
        return pages

    def create_page(self, properties: dict[str, Any]) -> dict[str, Any]:
        body = {
            "parent": {"database_id": self.database_id},
            "properties": properties,
        }
        return self._request("POST", "/pages", json=body)

    def update_page(self, page_id: str, properties: dict[str, Any],
                    archived: bool | None = None) -> dict[str, Any]:
        """PATCH a Notion page. If `archived` is given, the call also (un)archives it."""
        body: dict[str, Any] = {"properties": properties}
        if archived is not None:
            body["archived"] = archived
        return self._request("PATCH", f"/pages/{page_id}", json=body)

    def archive_page(self, page_id: str) -> None:
        self._request("PATCH", f"/pages/{page_id}", json={"archived": True})


# ---------------------------------------------------------------------------
# Local state cache (SQLite, stdlib only)
# ---------------------------------------------------------------------------

class StateCache:
    """Persistent local cache so `sync` can diff and skip unchanged rows.

    Stored as a single-table SQLite DB at ``sync/state.db``. Each row records the
    most recent state of a Notion page that we've successfully written. We use it
    as the source of truth for "does this Migaku key already exist in Notion?" and
    "did anything tracked on it change since last sync?". The cache row is written
    immediately after the matching Notion call returns so a kill -9 mid-run leaves
    Notion and the cache in sync.
    """

    SCHEMA_SQL = """
    CREATE TABLE IF NOT EXISTS words (
        migaku_key      TEXT PRIMARY KEY,
        page_id         TEXT NOT NULL,
        lang            TEXT,
        dict_form       TEXT,
        secondary       TEXT,
        known_status    TEXT,
        fail_rate       REAL,
        total_reviews   INTEGER,
        failed_reviews  INTEGER,
        part_of_speech  TEXT,
        last_synced     TEXT,
        archived        INTEGER NOT NULL DEFAULT 0,
        pinyin_marks    TEXT,
        pinyin_numeric  TEXT,
        sense_index     TEXT
    );
    CREATE INDEX IF NOT EXISTS idx_words_lang ON words(lang);
    """

    def __init__(self, path: Path) -> None:
        self.path = path
        self.conn = sqlite3.connect(str(path))
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA journal_mode=WAL;")
        self.conn.execute("PRAGMA synchronous=NORMAL;")
        with self.conn:
            self.conn.executescript(self.SCHEMA_SQL)
        # Idempotent migrations for older state.db files. Each ALTER is wrapped in its
        # own transaction so a partial migration on an interrupted run still progresses.
        cols = {row["name"] for row in self.conn.execute("PRAGMA table_info(words)")}
        if "pinyin" in cols and "pinyin_marks" not in cols:
            # The pre-D-schema cache used a single `pinyin` column. Rename it so we can
            # treat it as the new tone-marks column without losing data.
            with self.conn:
                self.conn.execute("ALTER TABLE words RENAME COLUMN pinyin TO pinyin_marks")
            cols.add("pinyin_marks")
            cols.discard("pinyin")
        for col in ("pinyin_marks", "pinyin_numeric", "sense_index"):
            if col not in cols:
                with self.conn:
                    self.conn.execute(f"ALTER TABLE words ADD COLUMN {col} TEXT")

    def close(self) -> None:
        self.conn.close()

    def load_all(self) -> dict[str, CachedRow]:
        out: dict[str, CachedRow] = {}
        for r in self.conn.execute("SELECT * FROM words"):
            out[r["migaku_key"]] = CachedRow(
                migaku_key=r["migaku_key"],
                page_id=r["page_id"],
                lang=r["lang"],
                dict_form=r["dict_form"],
                secondary=r["secondary"],
                known_status=r["known_status"],
                fail_rate=r["fail_rate"],
                total_reviews=r["total_reviews"],
                failed_reviews=r["failed_reviews"],
                part_of_speech=r["part_of_speech"],
                last_synced=r["last_synced"],
                archived=bool(r["archived"]),
                pinyin_marks=r["pinyin_marks"],
                pinyin_numeric=r["pinyin_numeric"],
                sense_index=r["sense_index"],
            )
        return out

    def upsert(self, row: CachedRow) -> None:
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO words (migaku_key, page_id, lang, dict_form, secondary,
                                   known_status, fail_rate, total_reviews, failed_reviews,
                                   part_of_speech, last_synced, archived,
                                   pinyin_marks, pinyin_numeric, sense_index)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(migaku_key) DO UPDATE SET
                    page_id        = excluded.page_id,
                    lang           = excluded.lang,
                    dict_form      = excluded.dict_form,
                    secondary      = excluded.secondary,
                    known_status   = excluded.known_status,
                    fail_rate      = excluded.fail_rate,
                    total_reviews  = excluded.total_reviews,
                    failed_reviews = excluded.failed_reviews,
                    part_of_speech = excluded.part_of_speech,
                    last_synced    = excluded.last_synced,
                    archived       = excluded.archived,
                    pinyin_marks   = excluded.pinyin_marks,
                    pinyin_numeric = excluded.pinyin_numeric,
                    sense_index    = excluded.sense_index
                """,
                (row.migaku_key, row.page_id, row.lang, row.dict_form, row.secondary,
                 row.known_status, row.fail_rate, row.total_reviews, row.failed_reviews,
                 row.part_of_speech, row.last_synced, 1 if row.archived else 0,
                 row.pinyin_marks, row.pinyin_numeric, row.sense_index),
            )

    def mark_archived(self, key: str, archived: bool, last_synced: str | None) -> None:
        with self.conn:
            self.conn.execute(
                "UPDATE words SET archived = ?, last_synced = COALESCE(?, last_synced) "
                "WHERE migaku_key = ?",
                (1 if archived else 0, last_synced, key),
            )

    def stats(self) -> dict[str, Any]:
        row = self.conn.execute(
            "SELECT COUNT(*) AS total, "
            "SUM(CASE WHEN archived=1 THEN 1 ELSE 0 END) AS archived, "
            "MAX(last_synced) AS last_synced "
            "FROM words"
        ).fetchone()
        return {
            "total": row["total"] or 0,
            "archived": row["archived"] or 0,
            "last_synced": row["last_synced"],
        }


# ---------------------------------------------------------------------------
# Property extraction (used when bootstrapping the cache from Notion)
# ---------------------------------------------------------------------------

def _prop_text(prop: dict[str, Any] | None) -> str:
    if not prop:
        return ""
    if "title" in prop:
        return "".join(t.get("plain_text", "") for t in (prop.get("title") or []))
    if "rich_text" in prop:
        return "".join(t.get("plain_text", "") for t in (prop.get("rich_text") or []))
    if "select" in prop:
        sel = prop.get("select") or {}
        return sel.get("name") or ""
    return ""


def _prop_number(prop: dict[str, Any] | None) -> float | None:
    if not prop:
        return None
    return prop.get("number")


def _prop_date_start(prop: dict[str, Any] | None) -> str | None:
    if not prop:
        return None
    d = prop.get("date") or {}
    return d.get("start")


def cache_row_from_notion_page(page: dict[str, Any]) -> CachedRow | None:
    """Project a Notion page object into a CachedRow. Returns None if the page has no
    Migaku key (those rows are out of scope for the sync)."""
    props = page.get("properties", {}) or {}
    key = _prop_text(props.get("Migaku key"))
    if not key:
        return None
    parts = key.split("|", 2)
    if len(parts) != 3:
        return None
    lang, dict_form, secondary = parts
    total = _prop_number(props.get("Total reviews"))
    failed = _prop_number(props.get("Failed reviews"))
    return CachedRow(
        migaku_key=key,
        page_id=page["id"],
        lang=lang,
        dict_form=dict_form,
        secondary=secondary,
        known_status=_prop_text(props.get("Status")) or None,
        fail_rate=_prop_number(props.get("Fail rate %")),
        total_reviews=int(total) if total is not None else None,
        failed_reviews=int(failed) if failed is not None else None,
        part_of_speech=_prop_text(props.get("Part of speech")) or None,
        last_synced=_prop_date_start(props.get("Last synced")),
        archived=bool(page.get("archived", False)),
        pinyin_marks=_prop_text(props.get("Pinyin")) or None,
        pinyin_numeric=_prop_text(props.get("Pinyin (numeric)")) or None,
        sense_index=_prop_text(props.get("Sense #")) or None,
    )


def cache_row_from_word(word: Word, page_id: str, last_synced: str | None,
                        archived: bool = False) -> CachedRow:
    # For zh, the "Sense #" column mirrors migaku's `secondary` (the disambiguation index).
    # For ja, `secondary` is the kana reading and not relevant to Sense #.
    sense_index = word.secondary if word.language == "zh" else None
    return CachedRow(
        migaku_key=word.key,
        page_id=page_id,
        lang=word.language,
        dict_form=word.dict_form,
        secondary=word.secondary,
        known_status=word.known_status or None,
        fail_rate=round(word.fail_rate, 2) if word.fail_rate is not None else None,
        total_reviews=word.total_reviews,
        failed_reviews=word.failed_reviews,
        part_of_speech=word.part_of_speech or None,
        last_synced=last_synced,
        archived=archived,
        pinyin_marks=word.pinyin_marks or None,
        pinyin_numeric=word.pinyin_numeric or None,
        sense_index=sense_index,
    )


def _approx_eq(a: float | None, b: float | None) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) < 1e-6


def has_tracked_changes(word: Word, cached: CachedRow) -> bool:
    """Return True if any tracked field on `word` differs from `cached`.

    Tracked fields: known_status, fail_rate, total_reviews, failed_reviews,
    part_of_speech, pinyin_marks, pinyin_numeric, sense_index. Meaning is intentionally
    NOT tracked — it's owned by the user / Notion AI.
    """
    if (word.known_status or None) != cached.known_status:
        return True
    word_fail = round(word.fail_rate, 2) if word.fail_rate is not None else None
    if not _approx_eq(word_fail, cached.fail_rate):
        return True
    if word.total_reviews != cached.total_reviews:
        return True
    if word.failed_reviews != cached.failed_reviews:
        return True
    if (word.part_of_speech or None) != cached.part_of_speech:
        return True
    if (word.pinyin_marks or None) != cached.pinyin_marks:
        return True
    if (word.pinyin_numeric or None) != cached.pinyin_numeric:
        return True
    word_sense = word.secondary if word.language == "zh" else None
    if (word_sense or None) != cached.sense_index:
        return True
    return False


# ---------------------------------------------------------------------------
# Notion property builders (write side)
# ---------------------------------------------------------------------------

def _rich(text: str | None) -> list[dict[str, Any]]:
    if not text:
        return []
    return [{"type": "text", "text": {"content": text[:1900]}}]


def build_properties(word: Word, *, include_meaning: bool, now_iso: str) -> dict[str, Any]:
    """Build the Notion properties payload for a word.

    `include_meaning` should be True for new pages (sets Meaning to blank) and False
    for updates (so the user's AI-filled Meaning is preserved).

    Column conventions:
      - "Pinyin"           -> tone-marked pinyin (zh only)
      - "Pinyin (numeric)" -> numeric-tone pinyin (zh only)
      - "Sense #"          -> migaku's disambiguation index, e.g. "0" or "1" (zh only)
      - For non-zh, "Pinyin" gets the migaku `secondary` (kana for ja, etc.) and the
        other two are left blank.
    """
    if word.language == "zh":
        pinyin_main = word.pinyin_marks or ""
        pinyin_numeric = word.pinyin_numeric or ""
        sense = word.secondary or ""
    else:
        pinyin_main = word.secondary or ""
        pinyin_numeric = ""
        sense = ""

    props: dict[str, Any] = {
        "Word": {"title": _rich(word.dict_form)},
        "Pinyin": {"rich_text": _rich(pinyin_main)},
        "Pinyin (numeric)": {"rich_text": _rich(pinyin_numeric)},
        "Sense #": {"rich_text": _rich(sense)},
        "Status": {"select": {"name": word.known_status} if word.known_status else None},
        "Language": {"select": {"name": word.language}},
        "Last synced": {"date": {"start": now_iso}},
        "Migaku key": {"rich_text": _rich(word.key)},
    }
    if word.fail_rate is not None:
        props["Fail rate %"] = {"number": round(word.fail_rate, 2)}
    if word.total_reviews is not None:
        props["Total reviews"] = {"number": word.total_reviews}
    if word.failed_reviews is not None:
        props["Failed reviews"] = {"number": word.failed_reviews}
    if word.part_of_speech:
        props["Part of speech"] = {"rich_text": _rich(word.part_of_speech)}
    if include_meaning:
        props["Meaning"] = {"rich_text": []}
    return props


# ---------------------------------------------------------------------------
# Sync orchestration
# ---------------------------------------------------------------------------

def merge_difficulty(words: list[Word], difficult: list[dict[str, Any]]) -> None:
    """Mutate `words` in-place, attaching fail-rate stats from /api/v1/words/difficult."""
    by_pair: dict[tuple[str, str], dict[str, Any]] = {}
    for d in difficult:
        by_pair[(d.get("dictForm", ""), d.get("secondary", ""))] = d
    for w in words:
        match = by_pair.get((w.dict_form, w.secondary))
        if not match:
            continue
        w.fail_rate = match.get("fail_rate")
        w.total_reviews = match.get("total_reviews")
        w.failed_reviews = match.get("failed_reviews")
        if not w.part_of_speech:
            w.part_of_speech = match.get("partOfSpeech") or None


def _bootstrap_cache_from_notion(notion: NotionClient,
                                 cache: StateCache | None) -> dict[str, CachedRow]:
    """Pull every page in the Notion DB and project them into a {key: CachedRow} dict.

    If `cache` is provided, also persist each row to disk (used on first real sync /
    `rebuild-cache`). If `cache` is None, the result is in-memory only (used by
    `--dry-run` when state.db doesn't yet exist).
    """
    pages = notion.query_all_pages()
    out: dict[str, CachedRow] = {}
    skipped = 0
    for page in pages:
        row = cache_row_from_notion_page(page)
        if row is None:
            skipped += 1
            continue
        if cache is not None:
            cache.upsert(row)
        out[row.migaku_key] = row
    log.info("Bootstrap: %d Notion pages -> %d cached rows (%d skipped — no Migaku key)",
             len(pages), len(out), skipped)
    return out


def run_sync(args: argparse.Namespace) -> int:
    migoku_url = os.getenv("MIGOKU_URL", "http://localhost:8080")
    api_key = os.getenv("MIGOKU_API_KEY")
    email = os.getenv("MIGOKU_EMAIL")
    password = os.getenv("MIGOKU_PASSWORD")

    notion_token = os.getenv("NOTION_TOKEN")
    notion_db = os.getenv("NOTION_DATABASE_ID")

    if not notion_token or not notion_db:
        log.error("NOTION_TOKEN and NOTION_DATABASE_ID must be set in .env")
        return 2

    migoku = MigokuClient(migoku_url, api_key)
    if not migoku.api_key:
        if not (email and password):
            log.error("No MIGOKU_API_KEY and no MIGOKU_EMAIL / MIGOKU_PASSWORD; cannot authenticate.")
            return 2
        migoku.login(email, password)

    statuses = [s.strip().upper() for s in (args.status or "").split(",") if s.strip()]
    if statuses == ["ALL"]:
        statuses = []
    log.info("Fetching words for lang=%s, statuses=%s ...",
             args.lang, statuses or "ALL")
    raw_words = list(migoku.list_words(args.lang, page_size=500, statuses=statuses or None))
    # Migoku has a pagination quirk where the last page can return overlapping rows;
    # dedupe by composite key (dictForm + secondary) so we never push duplicates to Notion.
    words: list[Word] = []
    seen: set[str] = set()
    for w in raw_words:
        if w.key in seen:
            continue
        seen.add(w.key)
        words.append(w)
    if len(words) != len(raw_words):
        log.info("Got %d words from migoku (%d unique after dedup)",
                 len(raw_words), len(words))
    else:
        log.info("Got %d words from migoku", len(words))

    if not words:
        log.warning("Migoku returned 0 words. Either there are none for lang=%s, "
                    "or the SRS database hasn't synced yet. Aborting.", args.lang)
        return 1

    # For Mandarin: derive real pinyin from each Hanzi (migoku's `secondary` is just a
    # disambiguation index, not a reading). We always generate both tone-marked and
    # numeric forms and write them to separate Notion columns.
    if args.lang == "zh":
        if not _PINYIN_AVAILABLE:
            log.error("Mandarin sync requires pypinyin. Install with `pip install pypinyin`.")
            return 2
        log.info("Generating pinyin (tone marks + numeric) for %d Hanzi ...", len(words))
        for w in words:
            w.pinyin_marks = compute_pinyin_marks(w.dict_form)
            w.pinyin_numeric = compute_pinyin_numeric(w.dict_form)

    diff_limit = int(os.getenv("SYNC_DIFFICULT_LIMIT", "2000"))
    log.info("Fetching difficult-words (limit=%d) for fail-rate enrichment ...", diff_limit)
    try:
        difficult = migoku.list_difficult_words(args.lang, diff_limit)
        log.info("Got %d difficult-word entries", len(difficult))
        merge_difficulty(words, difficult)
    except Exception as exc:  # difficulty enrichment is optional
        log.warning("Skipping difficulty enrichment: %s", exc)

    notion = NotionClient(notion_token, notion_db)

    # ---- Load (or bootstrap) the local cache ------------------------------
    cache: StateCache | None
    cache_rows: dict[str, CachedRow]
    if args.dry_run:
        cache = None
        if STATE_DB_PATH.exists():
            ro = StateCache(STATE_DB_PATH)
            cache_rows = ro.load_all()
            ro.close()
            log.info("Dry-run: loaded %d rows from local cache (no writes will occur)",
                     len(cache_rows))
        else:
            log.info("Dry-run: %s missing — bootstrapping in-memory cache from Notion ...",
                     STATE_DB_PATH.name)
            cache_rows = _bootstrap_cache_from_notion(notion, None)
    else:
        cache = StateCache(STATE_DB_PATH)
        cache_rows = cache.load_all()
        if not cache_rows:
            log.info("Local cache empty — bootstrapping from Notion (one-time, "
                     "preserves existing rows) ...")
            cache_rows = _bootstrap_cache_from_notion(notion, cache)

    now_iso = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    created = updated = unchanged = archived_count = 0
    seen_keys: set[str] = set()

    for i, word in enumerate(words, 1):
        if not word.dict_form:
            continue
        seen_keys.add(word.key)
        cached = cache_rows.get(word.key)

        if cached is None:
            # New word — CREATE in Notion, then write the cache row immediately.
            if args.dry_run:
                created += 1
            else:
                page = notion.create_page(
                    build_properties(word, include_meaning=True, now_iso=now_iso))
                page_id = page.get("id")
                if not page_id:
                    log.error("Notion create returned no page id for %s; aborting.", word.key)
                    return 3
                row = cache_row_from_word(word, page_id, now_iso)
                try:
                    cache.upsert(row)  # type: ignore[union-attr]
                except Exception as exc:
                    log.error(
                        "CRITICAL: Notion CREATE succeeded for %s (page %s) but cache write "
                        "FAILED (%s). Aborting before duplicate Notion rows pile up. "
                        "After fixing the cause, run `python sync.py rebuild-cache` "
                        "to re-sync state.db from Notion.",
                        word.key, page_id, exc,
                    )
                    return 3
                cache_rows[word.key] = row
                created += 1

        elif cached.archived:
            # Word is back after being archived — un-archive in Notion + push fresh state.
            if args.dry_run:
                updated += 1
            else:
                notion.update_page(
                    cached.page_id,
                    build_properties(word, include_meaning=False, now_iso=now_iso),
                    archived=False,
                )
                row = cache_row_from_word(word, cached.page_id, now_iso, archived=False)
                cache.upsert(row)  # type: ignore[union-attr]
                cache_rows[word.key] = row
                updated += 1

        elif has_tracked_changes(word, cached):
            if args.dry_run:
                updated += 1
            else:
                notion.update_page(
                    cached.page_id,
                    build_properties(word, include_meaning=False, now_iso=now_iso),
                )
                row = cache_row_from_word(word, cached.page_id, now_iso)
                cache.upsert(row)  # type: ignore[union-attr]
                cache_rows[word.key] = row
                updated += 1

        else:
            # Identical to cache — no Notion call, no last_synced bump.
            unchanged += 1

        if i % 200 == 0:
            log.info("  ... %d/%d processed (created=%d updated=%d unchanged=%d)",
                     i, len(words), created, updated, unchanged)

    # ---- Archive stale rows (still in cache, no longer in migoku for this lang) -----
    if args.archive_stale:
        stale = [r for r in cache_rows.values()
                 if r.migaku_key not in seen_keys
                 and r.lang == args.lang
                 and not r.archived]
        log.info("Archiving %d stale rows (no longer in migoku for lang=%s)",
                 len(stale), args.lang)
        for cached in stale:
            log.info("  archive %s", cached.migaku_key)
            if args.dry_run:
                archived_count += 1
                continue
            notion.archive_page(cached.page_id)
            cache.mark_archived(cached.migaku_key, archived=True, last_synced=now_iso)  # type: ignore[union-attr]
            archived_count += 1

    if cache is not None:
        cache.close()

    log.info(
        "Done. created=%d updated=%d unchanged=%d archived=%d "
        "(skipped=%d Notion calls, dry_run=%s)",
        created, updated, unchanged, archived_count, unchanged, args.dry_run,
    )
    return 0


def run_rebuild_cache(args: argparse.Namespace) -> int:
    """Drop sync/state.db and rebuild it from a fresh Notion query (no Notion writes)."""
    notion_token = os.getenv("NOTION_TOKEN")
    notion_db = os.getenv("NOTION_DATABASE_ID")
    if not notion_token or not notion_db:
        log.error("NOTION_TOKEN and NOTION_DATABASE_ID must be set in .env")
        return 2

    notion = NotionClient(notion_token, notion_db)
    log.info("Rebuilding %s from Notion (read-only — no Notion writes will occur) ...",
             STATE_DB_PATH)

    # Remove the old DB and any WAL/SHM/journal sidecars before recreating.
    for suffix in ("", "-wal", "-shm", "-journal"):
        p = Path(str(STATE_DB_PATH) + suffix)
        if p.exists():
            log.info("  removing %s", p.name)
            p.unlink()

    cache = StateCache(STATE_DB_PATH)
    rows = _bootstrap_cache_from_notion(notion, cache)
    cache.close()
    log.info("Rebuilt cache at %s with %d rows.", STATE_DB_PATH, len(rows))
    return 0


def run_login(args: argparse.Namespace) -> int:
    email = args.email or os.getenv("MIGOKU_EMAIL")
    password = args.password or os.getenv("MIGOKU_PASSWORD")
    if not (email and password):
        log.error("Provide --email and --password, or set MIGOKU_EMAIL / MIGOKU_PASSWORD in .env")
        return 2
    migoku_url = os.getenv("MIGOKU_URL", "http://localhost:8080")
    client = MigokuClient(migoku_url)
    key = client.login(email, password)
    print(f"\nMIGOKU_API_KEY={key}\n")
    print("Save that line in your .env to skip the login step on future syncs.")
    return 0


def run_status(args: argparse.Namespace) -> int:
    migoku_url = os.getenv("MIGOKU_URL", "http://localhost:8080")
    rc = 0
    try:
        r = requests.get(f"{migoku_url}/dev/status", timeout=5)
        r.raise_for_status()
        print(f"migoku at {migoku_url}: {r.json()}")
    except Exception as exc:
        print(f"migoku at {migoku_url}: NOT REACHABLE ({exc})")
        rc = 1

    api_key = os.getenv("MIGOKU_API_KEY")
    if api_key:
        try:
            r = requests.get(f"{migoku_url}/api/v1/stats/words",
                             params={"lang": args.lang},
                             headers={"X-Api-Key": api_key}, timeout=15)
            r.raise_for_status()
            print(f"stats/words [{args.lang}]: {r.json()}")
        except Exception as exc:
            print(f"stats/words: failed ({exc})")
            rc = 1
    else:
        print("No MIGOKU_API_KEY in .env — run `python sync.py login` first.")

    # Local cache stats.
    if STATE_DB_PATH.exists():
        cache = StateCache(STATE_DB_PATH)
        s = cache.stats()
        cache.close()
        print(f"local cache ({STATE_DB_PATH.name}):")
        print(f"  total cached rows:   {s['total']}")
        print(f"  archived rows:       {s['archived']}")
        print(f"  newest last_synced:  {s['last_synced'] or '(never)'}")
    else:
        print(f"local cache ({STATE_DB_PATH.name}): not initialised — "
              "first `sync` will create it (or run `rebuild-cache`)")

    return rc


# ---------------------------------------------------------------------------
# Interactive setup wizard
# ---------------------------------------------------------------------------

ENV_PATH = Path(__file__).resolve().parent / ".env"

NOTION_DB_TITLE_DEFAULT = "Migaku Vocab"
NOTION_DB_DESCRIPTION = (
    "Words synced from Migaku via the migoku API. The Meaning column is meant "
    "to be filled in by you / Notion AI; do not edit other columns as they "
    "will be overwritten on each sync."
)

# JSON properties payload for POST /v1/databases. Mirrors what notion-create-database
# produces in the project setup we used originally.
NOTION_DB_PROPERTIES: dict[str, Any] = {
    "Word":             {"title": {}},
    "Pinyin":           {"rich_text": {}},
    "Meaning":          {"rich_text": {}},
    "Pinyin (numeric)": {"rich_text": {}},
    "Status": {
        "select": {
            "options": [
                {"name": "KNOWN",    "color": "green"},
                {"name": "LEARNING", "color": "yellow"},
                {"name": "UNKNOWN",  "color": "gray"},
                {"name": "TRACKED",  "color": "blue"},
                {"name": "IGNORED",  "color": "red"},
            ]
        }
    },
    "Fail rate %":      {"number": {"format": "number"}},
    "Total reviews":    {"number": {"format": "number"}},
    "Failed reviews":   {"number": {"format": "number"}},
    "Part of speech":   {"rich_text": {}},
    "Language": {
        "select": {
            "options": [
                {"name": "zh", "color": "orange"},
                {"name": "ja", "color": "blue"},
                {"name": "en", "color": "purple"},
                {"name": "es", "color": "yellow"},
            ]
        }
    },
    "Last synced":      {"date": {}},
    "Migaku key":       {"rich_text": {}},
    "Sense #":          {"rich_text": {}},
}


def _prompt(label: str, *, current: str | None = None, secret: bool = False,
            allow_blank: bool = False, default: str | None = None) -> str:
    """Prompt the user for a value. Re-prompt until they enter something.

    `current` is the existing value (shown as "(unchanged)" hint if non-empty).
    `secret` uses getpass so the input isn't echoed.
    `default` is offered if the user just presses enter.
    """
    while True:
        suffix = ""
        if current:
            suffix = " (press enter to keep current value)"
        elif default:
            suffix = f" [{default}]"
        if secret:
            value = getpass.getpass(f"  {label}{suffix}: ").strip()
        else:
            value = input(f"  {label}{suffix}: ").strip()
        if not value:
            if current:
                return current
            if default:
                return default
            if allow_blank:
                return ""
            print("    (required — please enter a value)")
            continue
        return value


def _extract_notion_page_id(raw: str) -> str | None:
    """Pull the 32-hex page id out of a Notion URL or accept it raw."""
    raw = raw.strip()
    # Match any 32 hex chars (with or without dashes).
    m = re.search(r"([0-9a-fA-F]{32})", raw.replace("-", ""))
    if m:
        return m.group(1).lower()
    return None


def _read_env_file() -> dict[str, str]:
    """Parse the existing .env into a dict, preserving everything we don't touch."""
    if not ENV_PATH.exists():
        return {}
    out: dict[str, str] = {}
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        out[key.strip()] = value.strip()
    return out


def _write_env_file(values: dict[str, str]) -> None:
    """Write the .env, taking the .env.example as a template so comments/order survive.

    Updated keys overwrite their corresponding lines; existing comments and
    structure stay intact. New keys not in the template are appended.
    """
    template_path = ENV_PATH.with_name(".env.example")
    if template_path.exists():
        lines = template_path.read_text(encoding="utf-8").splitlines()
    else:
        lines = []

    seen: set[str] = set()
    out_lines: list[str] = []
    for line in lines:
        if "=" in line and not line.lstrip().startswith("#"):
            key, _, _ = line.partition("=")
            key = key.strip()
            if key in values:
                out_lines.append(f"{key}={values[key]}")
                seen.add(key)
                continue
        out_lines.append(line)

    extra = [k for k in values if k not in seen]
    if extra:
        out_lines.append("")
        out_lines.append("# Added by `python sync.py setup`")
        for k in extra:
            out_lines.append(f"{k}={values[k]}")

    ENV_PATH.write_text("\n".join(out_lines) + "\n", encoding="utf-8")


def _create_notion_database(token: str, parent_page_id: str, *,
                             title: str = NOTION_DB_TITLE_DEFAULT) -> tuple[str, str]:
    """Create the Migaku Vocab database under the given parent page.

    Returns (database_id, database_url).
    """
    body = {
        "parent": {"type": "page_id", "page_id": parent_page_id},
        "title": [{"type": "text", "text": {"content": title}}],
        "description": [{"type": "text", "text": {"content": NOTION_DB_DESCRIPTION}}],
        "properties": NOTION_DB_PROPERTIES,
    }
    headers = {
        "Authorization": f"Bearer {token}",
        "Notion-Version": NOTION_VERSION,
        "Content-Type": "application/json",
    }
    resp = requests.post(f"{NOTION_API}/databases", headers=headers, json=body, timeout=30)
    if resp.status_code == 401:
        raise RuntimeError("Notion rejected the token (401). Double-check NOTION_TOKEN.")
    if resp.status_code == 404:
        raise RuntimeError(
            "Notion couldn't find the parent page (404). The integration probably isn't "
            "connected to it yet — open the page in Notion -> ... -> Connections -> "
            "Connect to -> pick your integration, then re-run setup."
        )
    if not resp.ok:
        raise RuntimeError(f"Notion POST /databases -> {resp.status_code}: {resp.text[:500]}")
    data = resp.json()
    return data["id"], data.get("url") or ""


def run_setup(args: argparse.Namespace) -> int:
    """Interactive first-run wizard. Idempotent — re-run safely.

    Skips prompts for values already set in `.env` unless --force is passed.
    Auto-creates the Notion database if NOTION_DATABASE_ID isn't set.
    """
    print()
    print("==============================================================")
    print("  Migaku-Notion setup wizard")
    print("==============================================================")
    print()
    print("This will walk you through the one-time configuration:")
    print("  1. Migaku login credentials")
    print("  2. Notion integration token")
    print("  3. Notion parent page where the database will live")
    print("  4. Auto-create the Migaku Vocab database with the right schema")
    print()
    print("Existing values in .env will be kept unless you pass --force.")
    print()

    existing = _read_env_file()
    if args.force:
        existing = {}

    # --- Section 1: Migaku creds --------------------------------------------
    print("--- 1. Migaku login ---")
    print("  These are your normal Migaku account credentials. They never")
    print("  leave your machine — migoku derives a deterministic API key")
    print("  from them via HMAC.")
    email = _prompt("Migaku email", current=existing.get("MIGAKU_EMAIL"))
    password = _prompt("Migaku password", current=existing.get("MIGAKU_PASSWORD"),
                       secret=True)
    print()

    # --- Section 2: Notion integration token --------------------------------
    print("--- 2. Notion integration ---")
    notion_token = existing.get("NOTION_TOKEN", "")
    if not notion_token:
        print("  You need a Notion 'internal integration' so this script can")
        print("  read/write your database. I'll open the integrations page in")
        print("  your browser. Create one (any name, any workspace, default")
        print("  capabilities are fine), then copy the 'Internal Integration")
        print("  Secret' and paste it here.")
        try:
            input("  Press enter to open https://www.notion.so/profile/integrations ... ")
            webbrowser.open("https://www.notion.so/profile/integrations")
        except KeyboardInterrupt:
            return 130
    notion_token = _prompt("Notion integration secret",
                           current=existing.get("NOTION_TOKEN"), secret=True)
    print()

    # --- Section 3: Parent page --------------------------------------------
    print("--- 3. Notion parent page ---")
    parent_id = ""
    if not existing.get("NOTION_DATABASE_ID"):
        print("  Decide where the Migaku Vocab database should live (e.g. a")
        print("  page called 'Mandarin' or 'Migaku Word List'). Then:")
        print("    a) Open that page in Notion")
        print("    b) Top-right ... -> Connections -> Connect to -> pick")
        print("       the integration you just created")
        print("    c) Copy the page URL (or just the page ID) and paste it here")
        while not parent_id:
            raw = _prompt("Notion parent page URL or ID")
            parent_id = _extract_notion_page_id(raw)
            if not parent_id:
                print("    (couldn't find a 32-hex page ID in that. try again.)")
        print()

    # --- Section 4: Create the database (or skip if it already exists) ------
    db_id = existing.get("NOTION_DATABASE_ID", "")
    db_url = ""
    if db_id and not args.force:
        print(f"--- 4. Notion database (already configured: {db_id}) ---")
        print("  Skipping creation — NOTION_DATABASE_ID already set.")
        print("  (Pass --force if you want to create a new one.)")
    else:
        print("--- 4. Creating the Migaku Vocab database in Notion ---")
        try:
            db_id, db_url = _create_notion_database(notion_token, parent_id)
            print(f"  Created: {db_id}")
            if db_url:
                print(f"  URL    : {db_url}")
        except RuntimeError as exc:
            print(f"  ERROR: {exc}")
            return 1
    print()

    # --- Write .env --------------------------------------------------------
    print("--- 5. Writing .env ---")
    new_env: dict[str, str] = {
        **existing,
        "MIGAKU_EMAIL": email,
        "MIGAKU_PASSWORD": password,
        "NOTION_TOKEN": notion_token,
        "NOTION_DATABASE_ID": db_id,
    }
    new_env.setdefault("MIGOKU_URL", "http://localhost:8080")
    new_env.setdefault("MIGAKU_API_KEY", "")
    new_env.setdefault("MIGOKU_API_KEY", "")
    new_env.setdefault("SYNC_LANG", "zh")
    new_env.setdefault("SYNC_STATUS", "KNOWN,LEARNING")
    new_env.setdefault("SYNC_DIFFICULT_LIMIT", "2000")
    _write_env_file(new_env)
    print(f"  Wrote {ENV_PATH}")
    print()

    print("==============================================================")
    print("  Setup complete.")
    print("==============================================================")
    print()
    print("Next:")
    print("  1. Make sure migoku is running:  cd ../migoku && docker compose up -d")
    print("  2. Verify connectivity:          python sync.py status")
    print("  3. Preview the sync:             python sync.py sync --dry-run")
    print("  4. Run the sync:                 python sync.py sync")
    print()
    return 0


# ---------------------------------------------------------------------------
# Export: state.db -> CSV / XLSX / Google Sheets
# ---------------------------------------------------------------------------

# Canonical column order for all exports. Maps display name -> CachedRow attribute name.
# (Kept in sync with NOTION_DB_PROPERTIES order. "Meaning" is filled lazily from Notion
# when --with-meaning is passed, since we don't track it locally.)
EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("Word",             "dict_form"),
    ("Pinyin",           "pinyin_marks"),
    ("Meaning",          "_meaning"),       # populated only with --with-meaning
    ("Pinyin (numeric)", "pinyin_numeric"),
    ("Status",           "known_status"),
    ("Fail rate %",      "fail_rate"),
    ("Total reviews",    "total_reviews"),
    ("Failed reviews",   "failed_reviews"),
    ("Part of speech",   "part_of_speech"),
    ("Language",         "lang"),
    ("Last synced",      "last_synced"),
    ("Migaku key",       "migaku_key"),
    ("Sense #",          "sense_index"),
]


def _row_value(row: CachedRow, attr: str, meanings: dict[str, str] | None) -> Any:
    if attr == "_meaning":
        return (meanings or {}).get(row.migaku_key, "")
    return getattr(row, attr, "")


def _fetch_meanings_from_notion(notion: NotionClient) -> dict[str, str]:
    """One full Notion query to pick up just the Meaning column, keyed by Migaku key.

    Used when --with-meaning is passed to export. Re-uses the same paginated
    query_all_pages() helper the sync flow already uses; ~5 seconds for 1500 rows.
    """
    log.info("Fetching Meaning column from Notion ...")
    pages = notion.query_all_pages()
    out: dict[str, str] = {}
    for page in pages:
        props = page.get("properties", {}) or {}
        key = _prop_text(props.get("Migaku key"))
        if not key:
            continue
        meaning = _prop_text(props.get("Meaning"))
        if meaning:
            out[key] = meaning
    log.info("Got %d meanings (out of %d Notion rows)", len(out), len(pages))
    return out


def _export_rows(rows: list[CachedRow], lang: str | None,
                 statuses: list[str] | None, include_archived: bool) -> list[CachedRow]:
    """Apply CLI filters to the cache rows. Returns a stable-sorted list."""
    out = []
    status_set = set(statuses) if statuses else None
    for r in rows:
        if not include_archived and r.archived:
            continue
        if lang and r.lang != lang:
            continue
        if status_set and (r.known_status or "") not in status_set:
            continue
        out.append(r)
    # Sort by Migaku key so exports are deterministic across runs.
    out.sort(key=lambda r: r.migaku_key)
    return out


def export_csv(path: Path, rows: list[CachedRow], meanings: dict[str, str] | None) -> None:
    import csv
    headers = [name for name, _ in EXPORT_COLUMNS]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for r in rows:
            writer.writerow([_row_value(r, attr, meanings) for _, attr in EXPORT_COLUMNS])
    log.info("Wrote CSV: %s (%d rows)", path, len(rows))


def export_xlsx(path: Path, rows: list[CachedRow], meanings: dict[str, str] | None) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError(
            "XLSX export requires `openpyxl`. Install with `pip install openpyxl` "
            "(or re-run `pip install -r requirements.txt`)."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Migaku Vocab"

    headers = [name for name, _ in EXPORT_COLUMNS]
    ws.append(headers)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="EEEEEE")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill

    for r in rows:
        ws.append([_row_value(r, attr, meanings) for _, attr in EXPORT_COLUMNS])

    # Frozen header + auto-filter for instant filtering when opened.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # Reasonable column widths (rough rules of thumb).
    widths = {
        "Word": 14, "Pinyin": 16, "Meaning": 50, "Pinyin (numeric)": 18,
        "Status": 11, "Fail rate %": 11, "Total reviews": 13,
        "Failed reviews": 14, "Part of speech": 16, "Language": 9,
        "Last synced": 22, "Migaku key": 26, "Sense #": 8,
    }
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(header, 14)

    wb.save(path)
    log.info("Wrote XLSX: %s (%d rows)", path, len(rows))


def run_export(args: argparse.Namespace) -> int:
    """Export the local cache to one or more spreadsheet formats."""
    if not (args.csv or args.xlsx):
        log.error("Pass at least one of --csv PATH or --xlsx PATH.")
        return 2

    if not STATE_DB_PATH.exists():
        log.error("Local cache (%s) not initialised. Run `python sync.py sync` or "
                  "`python sync.py rebuild-cache` first.", STATE_DB_PATH.name)
        return 1

    cache = StateCache(STATE_DB_PATH)
    all_rows = list(cache.load_all().values())
    cache.close()

    statuses = [s.strip().upper() for s in (args.status or "").split(",") if s.strip()] or None
    if statuses == ["ALL"]:
        statuses = None
    rows = _export_rows(all_rows, args.lang or None, statuses, args.include_archived)
    log.info("Exporting %d rows (filtered from %d cached, lang=%s, status=%s, archived=%s)",
             len(rows), len(all_rows), args.lang or "ALL",
             ",".join(statuses) if statuses else "ALL", args.include_archived)

    meanings: dict[str, str] | None = None
    if args.with_meaning:
        notion_token = os.getenv("NOTION_TOKEN")
        notion_db = os.getenv("NOTION_DATABASE_ID")
        if not (notion_token and notion_db):
            log.error("--with-meaning requires NOTION_TOKEN and NOTION_DATABASE_ID in .env")
            return 2
        notion = NotionClient(notion_token, notion_db)
        meanings = _fetch_meanings_from_notion(notion)

    if args.csv:
        export_csv(Path(args.csv), rows, meanings)
    if args.xlsx:
        export_xlsx(Path(args.xlsx), rows, meanings)

    return 0


# ---------------------------------------------------------------------------
# Stats: unique-character count
# ---------------------------------------------------------------------------

def _is_cjk(ch: str) -> bool:
    """True if `ch` is a CJK ideograph (Han character).

    Covers the main blocks where >99% of practical Mandarin characters live:
    CJK Unified Ideographs (U+4E00-U+9FFF), Extension A (U+3400-U+4DBF), and
    Extension B (U+20000-U+2A6DF). Excludes punctuation, latin, digits.
    """
    if not ch:
        return False
    cp = ord(ch)
    return (0x4E00 <= cp <= 0x9FFF
            or 0x3400 <= cp <= 0x4DBF
            or 0x20000 <= cp <= 0x2A6DF)


def run_chars(args: argparse.Namespace) -> int:
    """Report unique Hanzi character counts from the local cache.

    Useful for tracking HSK preparation progress: a "word" in Migaku is often a
    multi-character compound, but the practical literacy metric is unique characters.
    """
    if not STATE_DB_PATH.exists():
        log.error("Local cache (%s) not initialised. Run `python sync.py sync` or "
                  "`python sync.py rebuild-cache` first.", STATE_DB_PATH.name)
        return 1

    cache = StateCache(STATE_DB_PATH)
    rows = cache.load_all()
    cache.close()

    # Group dict_form entries by status so the user can see the breakdown.
    by_status: dict[str, list[str]] = {}
    for row in rows.values():
        if row.lang != args.lang or row.archived:
            continue
        status = row.known_status or "(no status)"
        by_status.setdefault(status, []).append(row.dict_form)

    if not by_status:
        log.warning("No %s words in the local cache. Run `sync` first.", args.lang)
        return 1

    def chars_in(words: list[str]) -> set[str]:
        out: set[str] = set()
        for w in words:
            for ch in w:
                if _is_cjk(ch):
                    out.add(ch)
        return out

    print(f"\nUnique Hanzi character counts (lang={args.lang}, from local cache)")
    print("-" * 64)
    print(f"  {'Status':<14} {'Words':>8}  {'Unique chars':>14}")
    print(f"  {'-' * 14} {'-' * 8}  {'-' * 14}")
    cumulative_words = 0
    cumulative_chars: set[str] = set()
    for status in ("KNOWN", "LEARNING", "TRACKED", "UNKNOWN", "IGNORED"):
        words = by_status.get(status, [])
        if not words:
            continue
        chars = chars_in(words)
        print(f"  {status:<14} {len(words):>8}  {len(chars):>14}")
        cumulative_words += len(words)
        cumulative_chars |= chars

    # Useful "comprehension" subtotals.
    known = chars_in(by_status.get("KNOWN", []))
    known_plus_learning = chars_in(by_status.get("KNOWN", []) + by_status.get("LEARNING", []))
    print()
    print(f"  KNOWN only             : {len(known):>5} unique chars")
    print(f"  KNOWN + LEARNING       : {len(known_plus_learning):>5} unique chars")
    print(f"  All statuses combined  : {len(cumulative_chars):>5} unique chars "
          f"(across {cumulative_words} words)")

    if args.list:
        print("\nKNOWN + LEARNING characters:")
        # 30 per line for compact output
        chars_sorted = sorted(known_plus_learning)
        for i in range(0, len(chars_sorted), 30):
            print("  " + "".join(chars_sorted[i:i+30]))

    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--verbose", "-v", action="store_true")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_sync = sub.add_parser(
        "sync",
        help="Sync migoku words into Notion. Diffs against sync/state.db so re-runs "
             "only PATCH rows whose tracked fields actually changed.",
    )
    p_sync.add_argument("--lang", default=DEFAULT_LANG,
                        help=f"Migaku language code (default: {DEFAULT_LANG})")
    p_sync.add_argument("--status",
                        default=os.getenv("SYNC_STATUS", "KNOWN,LEARNING"),
                        help="Comma-separated migaku statuses to include "
                             "(KNOWN, LEARNING, UNKNOWN, IGNORED). "
                             "Use 'ALL' or empty to include everything. "
                             "Default: KNOWN,LEARNING")
    p_sync.add_argument("--dry-run", action="store_true",
                        help="Don't write to Notion or state.db; just log what would happen.")
    p_sync.add_argument("--archive-stale", action="store_true",
                        help="Archive Notion rows that no longer exist in migoku "
                             "for this lang+status filter")
    p_sync.set_defaults(func=run_sync)

    p_rebuild = sub.add_parser(
        "rebuild-cache",
        help="Delete sync/state.db and rebuild it from a fresh Notion query. "
             "Read-only — no Notion writes. Use after manual edits in the Notion UI "
             "or if state.db gets corrupted.",
    )
    p_rebuild.set_defaults(func=run_rebuild_cache)

    p_login = sub.add_parser("login", help="Derive a migoku API key and print it")
    p_login.add_argument("--email")
    p_login.add_argument("--password")
    p_login.set_defaults(func=run_login)

    p_status = sub.add_parser(
        "status",
        help="Show migoku connectivity, word counts, and local-cache stats.",
    )
    p_status.add_argument("--lang", default=DEFAULT_LANG)
    p_status.set_defaults(func=run_status)

    p_chars = sub.add_parser(
        "chars",
        help="Report unique Hanzi character counts from the local cache, "
             "broken down by Migaku status. Useful for tracking HSK progress.",
    )
    p_chars.add_argument("--lang", default=DEFAULT_LANG)
    p_chars.add_argument("--list", action="store_true",
                         help="Also print the full sorted list of KNOWN+LEARNING chars.")
    p_chars.set_defaults(func=run_chars)

    p_setup = sub.add_parser(
        "setup",
        help="Interactive first-run wizard. Walks through Migaku login, Notion "
             "integration setup, parent page selection, auto-creates the Migaku "
             "Vocab database with the right schema, and writes everything to .env.",
    )
    p_setup.add_argument("--force", action="store_true",
                         help="Re-prompt for every value, even ones already in .env. "
                              "Also creates a NEW Notion database, even if NOTION_DATABASE_ID "
                              "is already set.")
    p_setup.set_defaults(func=run_setup)

    p_export = sub.add_parser(
        "export",
        help="Export the local cache to CSV or XLSX. Reads from sync/state.db "
             "(no Notion API calls unless --with-meaning is set).",
    )
    p_export.add_argument("--csv", metavar="PATH",
                          help="Write a CSV file at PATH (UTF-8 with BOM, Excel-compatible).")
    p_export.add_argument("--xlsx", metavar="PATH",
                          help="Write an Excel workbook at PATH (frozen header + auto-filter). "
                               "Both Excel and Google Sheets (via File -> Import) accept it.")
    p_export.add_argument("--lang", default=DEFAULT_LANG,
                          help=f"Filter by language code (default: {DEFAULT_LANG}). "
                               "Pass empty to include all languages.")
    p_export.add_argument("--status",
                          default=os.getenv("SYNC_STATUS", "KNOWN,LEARNING"),
                          help="Comma-separated statuses to include, or ALL. "
                               "Default: KNOWN,LEARNING")
    p_export.add_argument("--include-archived", action="store_true",
                          help="Also include rows that have been archived in Notion.")
    p_export.add_argument("--with-meaning", action="store_true",
                          help="Also pull the Meaning column from Notion (one extra API "
                               "query, ~5 sec for 1500 rows). Without this flag, the "
                               "Meaning column in exports will be blank.")
    p_export.set_defaults(func=run_export)

    args = parser.parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
